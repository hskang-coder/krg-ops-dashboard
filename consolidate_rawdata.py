#!/usr/bin/env python3
"""
KRG 납부데이터 통합 스크립트
- Rawdata/ 폴더의 모든 납부데이터 xlsx를 통합
- 중복 제거: 계약번호 + 납부번호 복합키, 최신 파일(수정시각 기준) 데이터 유지
- 출력: Rawdata/통합_납부데이터.xlsx
- 로그: Rawdata/consolidation_log.txt
"""
import os
import sys
import warnings
from datetime import datetime

warnings.filterwarnings("ignore")

import openpyxl
from openpyxl import Workbook

BASE = os.path.dirname(os.path.abspath(__file__))
RAWDATA_DIR = os.path.join(BASE, "Rawdata")
OUTPUT_FILE = os.path.join(RAWDATA_DIR, "통합_납부데이터.xlsx")
LOG_FILE = os.path.join(RAWDATA_DIR, "consolidation_log.txt")

HEADER_KEYS = ("계약번호", "납부번호")  # 헤더 행 식별 및 복합키 컬럼


def log(msg, lines):
    line = f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {msg}"
    print(line)
    lines.append(line)


def find_header_row(ws):
    """상위 5행 내에서 계약번호/납부번호가 모두 포함된 행을 헤더로 판단."""
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
        vals = [str(v).strip() if v is not None else "" for v in row]
        if all(k in vals for k in HEADER_KEYS):
            return i, vals
    return None, None


def main():
    lines = []
    log("=== 납부데이터 통합 시작 ===", lines)

    if not os.path.isdir(RAWDATA_DIR):
        log(f"오류: Rawdata 폴더 없음: {RAWDATA_DIR}", lines)
        sys.exit(1)

    files = sorted(
        [
            f for f in os.listdir(RAWDATA_DIR)
            if f.lower().endswith(".xlsx")
            and not f.startswith("~$")
            and f != os.path.basename(OUTPUT_FILE)
        ],
        key=lambda f: os.path.getmtime(os.path.join(RAWDATA_DIR, f)),
    )

    if not files:
        log("통합할 xlsx 파일이 없습니다.", lines)
        write_log(lines)
        sys.exit(0)

    master_header = None
    records = {}  # (계약번호, 납부번호) -> row values
    total_read = 0

    for fname in files:  # 오래된 파일 -> 최신 파일 순, 최신 데이터가 덮어씀
        fpath = os.path.join(RAWDATA_DIR, fname)
        try:
            wb = openpyxl.load_workbook(fpath, data_only=True)
        except Exception as e:
            log(f"경고: {fname} 열기 실패 - {e}", lines)
            continue

        file_rows = 0
        for ws in wb.worksheets:
            hrow, header = find_header_row(ws)
            if hrow is None:
                continue
            if master_header is None:
                master_header = header
            idx_contract = header.index("계약번호")
            idx_payment = header.index("납부번호")

            for row in ws.iter_rows(min_row=hrow + 1, values_only=True):
                vals = list(row[: len(header)])
                if len(vals) < len(header):
                    vals += [None] * (len(header) - len(vals))
                c, p = vals[idx_contract], vals[idx_payment]
                if c is None and p is None:
                    continue

                def norm(v):
                    if v is None:
                        return ""
                    if isinstance(v, float) and v.is_integer():
                        return str(int(v))
                    return str(v).strip()

                key = (norm(c), norm(p))
                records[key] = vals  # 최신 파일이 덮어씀
                file_rows += 1
        wb.close()
        total_read += file_rows
        log(f"읽음: {fname} ({file_rows}행)", lines)

    if master_header is None:
        log("오류: 헤더(계약번호/납부번호)를 찾을 수 있는 파일이 없습니다.", lines)
        write_log(lines)
        sys.exit(1)

    out = Workbook()
    ws_out = out.active
    ws_out.title = "통합 납부데이터"
    ws_out.append(master_header)
    for vals in records.values():
        ws_out.append(vals)
    out.save(OUTPUT_FILE)

    dup_removed = total_read - len(records)
    log(f"파일 {len(files)}개, 총 {total_read}행 읽음 / 중복 제거 {dup_removed}행 / 최종 {len(records)}행", lines)
    log(f"저장 완료: {OUTPUT_FILE}", lines)
    log("=== 통합 완료 ===", lines)
    write_log(lines)


def write_log(lines):
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")


if __name__ == "__main__":
    main()
